"""Excel 读取"""

import os
from openpyxl import load_workbook
from .utils import ColumnInfo, TableData
from .type_parser import parse_marker_type


def read_excel_sheets(file_path: str, settings: dict) -> list:
    """读取 Excel 文件的所有 Sheet

    Args:
        file_path: Excel 文件路径
        settings: 配置字典

    Returns:
        list[TableData] 每个 Sheet 一个 TableData
    """
    comment_prefix = settings.get("comment_prefix", "#")
    header_row = settings.get("header_row", 1)
    marker_row = settings.get("marker_row", 2)
    type_row = settings.get("type_row", 3)
    data_start_row = settings.get("data_start_row", 5)

    # 获取文件名（不含扩展名）
    file_name = os.path.splitext(os.path.basename(file_path))[0]

    # 打开 workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)

    tables = []

    # 遍历所有 Sheet
    for sheet_name in wb.sheetnames:
        # 跳过 # 开头的 Sheet
        if sheet_name.startswith(comment_prefix):
            continue

        ws = wb[sheet_name]

        # 读取所有行
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < data_start_row:
            continue

        # 读取行1（字段名）
        header_row_data = all_rows[header_row - 1]

        # 读取行2（标记+类型组合）
        marker_row_data = all_rows[marker_row - 1]

        # 读取行3（类型定义，如果存在且行2没有类型信息）
        type_row_data = all_rows[type_row - 1] if type_row - 1 < len(all_rows) else None

        # 构建列信息
        columns = []
        for col_idx, (header, marker_cell) in enumerate(zip(header_row_data, marker_row_data)):
            # 跳过空列
            if header is None:
                continue

            header_str = str(header).strip()

            # 列注释：字段名以 # 开头跳过
            if header_str.startswith(comment_prefix):
                continue

            # 解析标记和类型
            marker_str = str(marker_cell).strip() if marker_cell else ""
            marker, type_def = parse_marker_type(marker_str)

            # 如果行2没有类型信息，从行3获取
            if type_def.type == "string" and marker_str in ("C", "S", "CS", ""):
                if type_row_data and col_idx < len(type_row_data):
                    type_cell = type_row_data[col_idx]
                    if type_cell:
                        type_str = str(type_cell).strip()
                        if type_str and type_str != header_str:
                            # 行3有独立的类型定义
                            from .type_parser import parse_type
                            type_def = parse_type(type_str)

            columns.append(ColumnInfo(
                name=header_str,
                marker=marker,
                type_def=type_def,
            ))

        # 读取数据行
        rows = []
        for row_idx, row_data in enumerate(all_rows[data_start_row - 1:], start=data_start_row):
            if not row_data or row_data[0] is None:
                continue

            # 行注释：第一列以 # 开头跳过
            first_cell = str(row_data[0]).strip() if row_data[0] else ""
            if first_cell.startswith(comment_prefix):
                continue

            # 构建行数据
            row_dict = {}
            for col_idx, col_info in enumerate(columns):
                # 找到该列在 Excel 中的位置
                excel_col_idx = _find_column_index(header_row_data, col_info.name, comment_prefix)
                if excel_col_idx is None or excel_col_idx >= len(row_data):
                    row_dict[col_info.name] = None
                else:
                    cell_value = row_data[excel_col_idx]
                    row_dict[col_info.name] = cell_value

            rows.append(row_dict)

        # 输出文件名 = Sheet名
        table_name = sheet_name

        tables.append(TableData(name=table_name, columns=columns, rows=rows))

    wb.close()
    return tables


def _find_column_index(header_row, column_name: str, comment_prefix: str = "#"):
    """在表头行中查找列名的索引"""
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if cell_str.startswith(comment_prefix):
            continue
        if cell_str == column_name:
            return idx
    return None
